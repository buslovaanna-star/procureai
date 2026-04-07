import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math, io, re
from datetime import date, timedelta

st.set_page_config(
    page_title="ProcureAI — Аналіз закупівель",
    page_icon="📦", layout="wide",
    initial_sidebar_state="expanded",
)

UA_MO = {'Січень':1,'Лютий':2,'Березень':3,'Квітень':4,'Травень':5,'Червень':6,
         'Липень':7,'Серпень':8,'Вересень':9,'Жовтень':10,'Листопад':11,'Грудень':12}
BUILT_IN_K = [1.1,1.15,1.18,1.05,0.95,1.34,1.34,0.9,0.92,1.2,1.5,1.56]

# ── Хелпери ──────────────────────────────────
def cs(v):
    """clean string — прибирає \xa0 та зайві пробіли"""
    return str(v).replace('\xa0', ' ').strip() if v is not None else ''

def sn(v):
    """safe number — конвертує рядок з пробілами і комами"""
    if v is None: return None
    try:
        f = float(cs(v).replace(' ', '').replace(',', '.'))
        return f
    except: return None

def parse_disc(v):
    """'Знижка 15 %', '15%', 0.15, 15 → 15.0"""
    if v is None: return 0
    s = cs(v)
    if not s or s in ('-', ''): return 0
    nums = re.findall(r'\d+[.,]?\d*', s)
    if not nums: return 0
    try:
        val = float(nums[0].replace(',', '.'))
        if val < 1: val *= 100
        return round(val, 1) if val > 0 else 0
    except: return 0

def mo_num(label):
    return UA_MO.get(cs(label).split()[0], 0)

def mo_year(label):
    for p in cs(label).split():
        if p.isdigit() and len(p) == 4: return int(p)
    return 0

# ── Завантаження Excel ────────────────────────
@st.cache_data(show_spinner=False)
def load_wb(file_bytes):
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

# ── Парсинг шаблону ───────────────────────────
def parse_template(wb):
    """
    Шаблон має 3 вкладки:
      'продажі дані'      — блоки по 5 колонок: Назва|Артикул|Кількість|Рентаб|_
      'наявність на складі' — Номенклатура|Артикул|Залишок|Замовлено
      'Залишки'           — блоки по 4 колонки: Артикул|Назва|Дні|_ (по місяцях)
    """
    errors = []

    # Знаходимо вкладки (нечутливо до регістру)
    sheets = {n.lower(): n for n in wb.sheetnames}
    def find(key):
        for k, v in sheets.items():
            if key in k: return wb[v]
        return None

    ws_sales = find('продаж')
    ws_stock = find('наявн')
    ws_avail = find('залиш')

    if ws_sales is None: errors.append("Не знайдено вкладку 'продажі дані'")
    if ws_stock is None: errors.append("Не знайдено вкладку 'наявність на складі'")
    if ws_avail is None: errors.append("Не знайдено вкладку 'Залишки'")
    if errors: return None, None, None, None, errors

    # ── Продажі ──
    all_rows = list(ws_sales.iter_rows(values_only=True))
    # Рядок 2 (idx 1) містить місяці "Період: Місяць YYYY р."
    row2 = all_rows[1] if len(all_rows) > 1 else []
    month_blocks = []
    for i, v in enumerate(row2):
        label = cs(v)
        if 'Період' in label and mo_num(label.replace('Період:', '').replace('Період :', '').strip().rstrip('.')) > 0:
            clean_label = label.replace('Період:', '').replace('Період :', '').strip().rstrip('.').strip()
            month_blocks.append((i, clean_label))  # 0-indexed col offset

    if not month_blocks:
        errors.append("Не знайдено місяців у вкладці 'продажі дані'")
        return None, None, None, None, errors

    months_labels = [lbl for _, lbl in month_blocks]

    # Дані з рядка 14 (idx 13)
    sku_data = {}
    DATA_START = 13
    for row in all_rows[DATA_START:]:
        for ci, label in month_blocks:
            sku_val  = row[ci+1] if ci+1 < len(row) else None
            name_val = row[ci]   if ci   < len(row) else None
            qty_val  = row[ci+2] if ci+2 < len(row) else None
            rent_val = row[ci+3] if ci+3 < len(row) else None

            if not sku_val: continue
            sku = cs(sku_val)
            if not sku or sku in ('Артикул', 'IHERB', '1*', ''): continue

            name = cs(name_val) if name_val else sku
            qty  = max(sn(qty_val) or 0, 0)
            rent = sn(rent_val)

            if sku not in sku_data:
                sku_data[sku] = {'name': name, 'months': {}}
            if label not in sku_data[sku]['months']:
                sku_data[sku]['months'][label] = [0.0, None]
            sku_data[sku]['months'][label][0] += qty
            if rent is not None and qty > 0:
                sku_data[sku]['months'][label][1] = rent

    # ── Наявність (залишок + в дорозі) ──
    stock_map = {}
    all_stock = list(ws_stock.iter_rows(values_only=True))
    # Рядок 2: заголовки, Рядок 4+: дані
    for row in all_stock[3:]:
        if len(row) < 2 or not row[1]: continue
        sku = cs(row[1])
        if not sku or sku == '1*': continue
        nm  = cs(row[0]) if row[0] else sku
        st  = max(sn(row[2]) or 0, 0)
        tr  = max(sn(row[3]) or 0, 0) if len(row) > 3 and row[3] else 0
        stock_map[sku] = (st, tr, nm)

    # ── Залишки (дні на складі по місяцях) ──
    avail_map = {}
    all_av = list(ws_avail.iter_rows(values_only=True))
    # Рядок 2 (idx 1): місяці, кожні 4 колонки: Артикул|Назва|Дні|_
    row2_av = all_av[1] if len(all_av) > 1 else []
    av_blocks = []
    for i, v in enumerate(row2_av):
        label = cs(v)
        if 'Період' in label:
            clean = label.replace('Період:', '').replace('Період :', '').strip().rstrip('.').strip()
            av_blocks.append((i, clean))

    for row in all_av[3:]:  # з рядка 4
        for ci, label in av_blocks:
            sku_val  = row[ci]   if ci   < len(row) else None
            days_val = row[ci+2] if ci+2 < len(row) else None
            if not sku_val: continue
            sku = cs(sku_val)
            if not sku or sku in ('Артикул', '1*', ''): continue
            days = max(sn(days_val) or 0, 0)
            if sku not in avail_map: avail_map[sku] = {}
            avail_map[sku][label] = avail_map[sku].get(label, 0) + days

    return sku_data, months_labels, stock_map, avail_map, []

# ── Парсинг файлу цін ─────────────────────────
def parse_prices(wb):
    """
    iHerb Catalog: 1 вкладка, рядок 1 = заголовки
    Артикул | Название | Цена | Стара ціна | Наличие | Скидка | Рейтинг | Відгуки | Продано за 30 дн | Ссылка
    """
    ws = wb[wb.sheetnames[0]]
    price_map = {}
    rows = list(ws.iter_rows(values_only=True))
    # Знаходимо рядок заголовків
    header_row = 0
    for i, row in enumerate(rows[:5]):
        cols = [cs(v).lower() for v in row if v]
        if any('артикул' in c or 'sku' in c for c in cols):
            header_row = i + 1
            break
    for row in rows[header_row:]:
        if not row[0]: continue
        sku = cs(row[0])
        if not sku: continue
        price     = sn(row[2]) if len(row) > 2 else None
        avail_str = cs(row[4]).lower() if len(row) > 4 and row[4] else ''
        in_stock  = 'наявн' in avail_str or avail_str in ('1', 'true', 'yes')
        disc      = parse_disc(row[5]) if len(row) > 5 else 0
        price_map[sku] = (price, in_stock, disc)
    return price_map

# ── Основний аналіз ───────────────────────────
def run_analysis(sku_data, months_labels, stock_map, avail_map, price_map, params):
    today     = date.today()
    CUR_MO    = today.month; CUR_YR = today.year; CUR_DAY = today.day
    cur_scale = 30.0 / CUR_DAY

    MG_MIN    = params['mg_min']; LEAD = params['lead']
    SAFETY    = params['safety']; SAFETY60 = params['safety60']
    SAFETY90  = params.get('safety90', 90)
    A_MULT    = params['a_mult']; MG_A = params['mg_a']
    MIN_MO    = params['min_months']; MIN_QTY = params['min_qty']
    AA        = params['avail_alpha']; LAM = params['lambda_val']
    LOW_AV    = params['low_avail']
    DISC_THR  = params['disc_thr']; DISC_THR2 = params.get('disc_thr2', 40)

    n = len(months_labels)

    def is_cur(lbl):  return mo_num(lbl) == CUR_MO and mo_year(lbl) == CUR_YR
    def is_done(lbl):
        mn, yr = mo_num(lbl), mo_year(lbl)
        return yr < CUR_YR or (yr == CUR_YR and mn < CUR_MO)

    mo_complete = [is_done(lbl) for lbl in months_labels]
    mo_is_cur   = [is_cur(lbl)  for lbl in months_labels]
    exp_w = [math.exp(LAM * (i - (n-1))) for i in range(n)]

    # Pass 1 — season K
    mo_cs = [0.0] * n; mo_cc = [0] * n
    for sku, sd in sku_data.items():
        for i, lbl in enumerate(months_labels):
            qty, rent = sd['months'].get(lbl, [0.0, None])
            inc = (rent >= MG_MIN) if rent is not None else (qty > 0)
            if inc and qty > 0 and mo_complete[i]:
                mo_cs[i] += qty; mo_cc[i] += 1

    ca_  = [mo_cs[i]/mo_cc[i] if mo_cc[i] > 0 else 0 for i in range(n)]
    cv   = [ca_[i] for i in range(n) if mo_complete[i] and ca_[i] > 0]
    global_avg = sum(cv)/len(cv) if cv else 1.0
    cur_ks = [ca_[i]/global_avg for i, lbl in enumerate(months_labels)
              if mo_complete[i] and mo_num(lbl) == CUR_MO and ca_[i] > 0]
    season_K = sum(cur_ks)/len(cur_ks) if cur_ks else BUILT_IN_K[CUR_MO-1]

    # Pass 2 — per SKU
    results = []; excl_mg = []; sporadic = []

    for sku, sd in sku_data.items():
        st_info = stock_map.get(sku)
        stock   = st_info[0] if st_info else 0
        transit = st_info[1] if st_info else 0
        nm      = st_info[2] if st_info and st_info[2] else sd['name']
        eff     = stock + transit
        av_mo   = avail_map.get(sku, {})

        # Зважений avg/день
        ws_ = 0.0; wd_ = 0.0; cd_clean = 0.0; rd = n * 30
        for i, lbl in enumerate(months_labels):
            qty, rent = sd['months'].get(lbl, [0.0, None])
            inc = (rent >= MG_MIN) if rent is not None else (qty > 0)
            if not inc: continue
            ad = av_mo.get(lbl, 30)
            wt = exp_w[i]
            q_ = qty * (cur_scale if mo_is_cur[i] else 1)
            a_ = min(ad  * (cur_scale if mo_is_cur[i] else 1), 30)
            if a_ > 0:
                ws_ += q_ * wt; wd_ += a_ * wt
            cd_clean += a_

        avg_day   = ws_ / wd_ if wd_ > 0 else 0
        avail_pct = round(cd_clean / rd * 100) if rd > 0 else 0
        avail_K   = math.pow(avail_pct/100, AA) if avail_pct > 0 else 0

        # Середня маржа (тільки чисті місяці)
        rents = [r for lbl in months_labels
                 for q, r in [sd['months'].get(lbl, [0, None])]
                 if r is not None and r >= MG_MIN]
        avg_margin = round(sum(rents)/len(rents), 1) if rents else None

        # Фільтр маржі
        if avg_margin is not None and avg_margin < MG_MIN:
            cmo = sum(1 for lbl in months_labels if sd['months'].get(lbl,[0,None])[0]>0)
            ts  = sum(sd['months'].get(lbl,[0,None])[0] for lbl in months_labels)
            excl_mg.append({'sku':sku,'name':nm,'avg_margin':avg_margin,
                             'clean_months':cmo,'total_sold':round(ts,1)})
            continue

        # Постійний попит
        cmo = sum(1 for lbl in months_labels if sd['months'].get(lbl,[0,None])[0]>0)
        ts  = sum(sd['months'].get(lbl,[0,None])[0] for lbl in months_labels)
        is_sp = not (cmo >= MIN_MO and ts >= MIN_QTY)
        reason = ''
        if is_sp:
            parts = []
            if cmo < MIN_MO:  parts.append(f"{cmo}<{MIN_MO}міс")
            if ts  < MIN_QTY: parts.append(f"{round(ts):.0f}<{MIN_QTY}шт")
            reason = ', '.join(parts)

        # ABC
        if avg_margin is None:    abc, am = '?', 1.0
        elif avg_margin >= MG_A:  abc, am = 'A', A_MULT
        else:                     abc, am = 'B', 1.0

        # Тренд
        clean_m = [(lbl, q) for lbl in months_labels
                   for q, r in [sd['months'].get(lbl,[0,None])]
                   if q > 0 and av_mo.get(lbl, 30) > 0]
        if len(clean_m) >= 6:
            f3 = sum(q for _,q in clean_m[:3])/3
            l3 = sum(q for _,q in clean_m[-3:])/3
            tr_r = round(l3/f3, 2) if f3 > 0 else None
            trend = ('↑ зростає' if tr_r and tr_r > 1.3 else
                     '↓ спадає'  if tr_r and tr_r < 0.7 else '→ стабільно')
        else:
            l3 = sum(q for _,q in clean_m[-3:])/3 if clean_m else 0
            trend = '↑ новий' if l3 > 0 else '—'

        # Ціни
        pi = price_map.get(sku)
        p_avail = pi[1] if pi else True
        p_disc  = pi[2] if pi else 0
        p_price = pi[0] if pi else None
        use_60  = not is_sp and p_avail and p_disc >= DISC_THR
        use_90  = not is_sp and p_avail and p_disc >= DISC_THR2
        safety_disc = SAFETY90 if use_90 else (SAFETY60 if use_60 else 0)

        # Маржинальний дохід
        mi_day = None
        if p_price and avg_margin and 0 < avg_margin < 100:
            sell_price = p_price / (1 - avg_margin/100)
            mi_day = round(avg_day * (sell_price - p_price), 4)

        dl = round(stock/avg_day) if avg_day > 0 else 999
        st2 = ('Критично' if dl<5 else 'Низько' if dl<15
               else 'Надлишок' if dl>90 else 'Норма')
        rec   = max(0,round(avg_day*avail_K*am*(LEAD+SAFETY  )*season_K-eff)) if not is_sp and p_avail else 0
        rec60 = max(0,round(avg_day*avail_K*am*(LEAD+safety_disc)*season_K-eff)) if safety_disc else 0
        zero_date = (today+timedelta(days=int(dl))).strftime('%d.%m.%Y') if 0<=dl<999 else None

        row_d = dict(
            sku=sku, name=nm, abc=abc, avg_margin=avg_margin,
            avail_pct=avail_pct, avail_K=round(avail_K,3),
            trend=trend, avg_day=round(avg_day,4),
            stock=stock, transit=transit, eff_stock=eff,
            days_left=dl, zero_date=zero_date, status=st2,
            is_sporadic=is_sp, sporadic_reason=reason,
            season_K=round(season_K,3), rec=rec, rec_60=rec60,
            use_60=use_60, use_90=use_90, safety_disc=safety_disc,
            price_disc=round(p_disc,1), buy_price=p_price,
            mi_day=mi_day, low_avail=avail_pct < LOW_AV,
        )
        if is_sp: sporadic.append(row_d)
        else:     results.append(row_d)

    # ABC по MI (Pareto 70/90)
    mi_vals = sorted([r['mi_day'] for r in results if r['mi_day']], reverse=True)
    total_mi = sum(mi_vals); cum = 0; ta = tb = None
    for v in mi_vals:
        cum += v
        if ta is None and cum >= total_mi*0.70: ta = v
        if tb is None and cum >= total_mi*0.90: tb = v
    for r in results:
        mi = r['mi_day']
        if mi is None:          r['abc_mi'] = '?'
        elif ta and mi >= ta:   r['abc_mi'] = 'A'
        elif tb and mi >= tb:   r['abc_mi'] = 'B'
        else:                   r['abc_mi'] = 'C'

    meta = dict(season_K=season_K, global_avg=global_avg, n_months=n,
                cur_day=CUR_DAY, cur_scale=cur_scale,
                total_mi=total_mi, ta=ta, tb=tb, months=months_labels)
    return dict(regular=results, sporadic=sporadic, excl_mg=excl_mg, meta=meta)

# ── Генерація Excel ───────────────────────────
def gen_excel(data, params):
    results = data['regular']; sporadic = data['sporadic']
    meta    = data['meta']; today = date.today()

    th = Side(style="thin", color="BFBFBF")
    def tb(): return Border(left=th,right=th,top=th,bottom=th)
    def fl(c): return PatternFill("solid",start_color=c,fgColor=c)
    def hf(color="FFFFFF",bold=True,sz=10): return Font(name="Arial",bold=bold,color=color,size=sz)
    def cf(bold=False,sz=10,color="000000"): return Font(name="Arial",bold=bold,size=sz,color=color)
    def ca(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def la(): return Alignment(horizontal="left",vertical="center")

    C = dict(main="1F4E79",red="C62828",amber="E65100",green="1B5E20",blue="0D47A1",
             gray="546E7A",orange="BF360C",teal="006064",gold="F57F17",
             red_l="FFEBEE",amber_l="FFF3E0",green_l="E8F5E9",blue_l="E3F2FD",
             gray_l="ECEFF1",orange_l="FBE9E7",gold_l="FFFDE7",low_l="FCE4EC",
             row1="EBF3FB",row2="F2F9EC",white="FFFFFF")
    ST = {'Критично':(C['red'],C['red_l']),'Низько':(C['amber'],C['amber_l']),
          'Норма':(C['green'],C['green_l']),'Надлишок':(C['blue'],C['blue_l'])}
    AB    = {'A':C['green'],'B':C['blue'],'?':C['gray']}
    AB_MI = {'A':C['gold'],'B':C['blue'],'C':C['gray'],'?':C['gray']}
    TR    = {'↑ зростає':C['green'],'↑ новий':C['teal'],'↓ спадає':C['red'],
             '→ стабільно':C['gray'],'—':C['gray']}

    wb = Workbook()
    p = (f"Маржа≥{params['mg_min']}% | Lead {params['lead']}д | Safety {params['safety']}д | "
         f"α={params['avail_alpha']} | λ={params['lambda_val']} | K={meta['season_K']:.3f}")

    def ws_init(ws, title, warn, cols, wc="7F3F00", wb_="FFF9C4"):
        lc = get_column_letter(len(cols))
        ws.merge_cells(f"A1:{lc}1"); ws["A1"] = title
        ws["A1"].font=hf(sz=9); ws["A1"].fill=fl(C['main'])
        ws["A1"].alignment=la(); ws.row_dimensions[1].height=20
        ws.merge_cells(f"A2:{lc}2"); ws["A2"] = warn
        ws["A2"].font=Font(name="Arial",size=9,italic=True,color=wc)
        ws["A2"].fill=fl(wb_); ws["A2"].alignment=la(); ws.row_dimensions[2].height=18
        for col,(h,w) in enumerate(cols,1):
            c=ws.cell(row=3,column=col,value=h)
            c.font=hf(sz=9); c.fill=fl(C['main']); c.alignment=ca(); c.border=tb()
            ws.column_dimensions[get_column_letter(col)].width=w
        ws.row_dimensions[3].height=40

    # Sheet 1 — Замовлення
    ws1 = wb.active; ws1.title = "Замовлення"
    COLS1 = [("SKU",13),("Назва",42),("ABC\nмарж%",7),("ABC\nMI",7),
             ("Маржа %",10),("MI/день",11),("Залишок+\nтранзит",10),
             ("Дата нуля",11),("Дні\nдо нуля",9),("Тренд",11),
             ("Замовити\n14+30д",12),("Замовити\n(знижка)",13)]
    ws_init(ws1, f"ЗАМОВЛЕННЯ | {today.strftime('%d.%m.%Y')} | {p}",
            f"⚑ = наявність <{params['low_avail']}%. "
            f"Золото ABC_MI=A (топ 70% маржинального доходу/день).", COLS1)

    order = sorted([r for r in results if r['rec']>0 or r['rec_60']>0],
                   key=lambda x: x['days_left'])
    for ri, r in enumerate(order):
        row = ri+4; dl = r['days_left']; low = r.get('low_avail', False)
        _, bg = ST.get(r['status'], (C['gray'],C['gray_l']))
        if low: bg = C['low_l']
        elif r['status']=='Норма': bg = C['row1'] if ri%2==0 else C['row2']
        tr = r.get('trend','—'); mi = r.get('mi_day'); abc_mi = r.get('abc_mi','?')
        sd_ = r.get('safety_disc',0)
        vals = [r['sku'],r['name'],r['abc'],abc_mi,r['avg_margin'],mi,r['eff_stock'],
                r.get('zero_date'),dl if dl<999 else None,tr,r['rec'],
                f"{r['rec_60']} шт ({sd_}д)" if r['rec_60']>0 else "—"]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=row,column=col,value=val); c.fill=fl(bg); c.border=tb()
            if col==1:
                c.value=("⚑ " if low else "")+str(r['sku'])
                c.font=Font(name="Arial",bold=True,size=10,color="880E4F" if low else "000000"); c.alignment=la()
            elif col==2: c.font=cf(sz=9); c.alignment=la()
            elif col==3: c.font=Font(name="Arial",bold=True,size=11,color=AB.get(r['abc'],C['gray'])); c.alignment=ca()
            elif col==4: c.font=Font(name="Arial",bold=True,size=11,color=AB_MI.get(abc_mi,C['gray'])); c.alignment=ca()
            elif col==5:
                if val: c.number_format='0.0"%"'
                c.font=Font(name="Arial",bold=True,size=10,color=C['green'] if (val or 0)>=params['mg_a'] else C['blue']); c.alignment=ca()
            elif col==6:
                if val: c.number_format='#,##0.00'
                ta_=meta.get('ta') or 0; tb__=meta.get('tb') or 0
                c.font=Font(name="Arial",bold=True,size=10,
                            color=C['gold'] if (val or 0)>=ta_ else C['blue'] if (val or 0)>=tb__ else C['gray']); c.alignment=ca()
            elif col==7: c.font=cf(bold=True,sz=10); c.alignment=ca()
            elif col==8:
                if val:
                    c.font=Font(name="Arial",bold=True,size=10,color=C['red'] if dl<5 else C['amber'] if dl<15 else "000000")
                else: c.value="∞"; c.font=cf(sz=10,color="AAAAAA")
                c.alignment=ca()
            elif col==9:
                c.font=Font(name="Arial",bold=True,size=11,
                            color=C['red'] if dl<5 else C['amber'] if dl<15 else C['green']); c.alignment=ca()
            elif col==10: c.font=Font(name="Arial",bold=True,size=10,color=TR.get(tr,C['gray'])); c.alignment=ca()
            elif col==11:
                if val and val>0: c.font=Font(name="Arial",bold=True,size=12,color=C['main'])
                else: c.value="—"; c.font=cf(sz=10,color="BBBBBB")
                c.alignment=ca()
            elif col==12:
                if r['rec_60']>0:
                    c.font=Font(name="Arial",bold=True,size=11,color=C['amber']); c.fill=fl(C['amber_l'])
                else: c.font=cf(sz=10,color="BBBBBB")
                c.alignment=ca()
    ws1.freeze_panes="A4"; ws1.auto_filter.ref=f"A3:{get_column_letter(len(COLS1))}{len(order)+3}"

    # Sheet 2 — Аналіз SKU
    ws2 = wb.create_sheet("Аналіз SKU")
    COLS2 = [("SKU",13),("Назва",42),("ABC\nмарж%",7),("ABC\nMI",7),
             ("Маржа %",10),("MI/день",11),("Наявн %",8),("Тренд",10),
             ("avg/день",10),("Залишок",8),("Транзит",8),
             ("Дата нуля",11),("Дні\nдо нуля",9),("Замовити\n14+30д",12),("Замовити\n(знижка)",12)]
    ws_init(ws2, f"Аналіз SKU — {len(results)} регулярних | {today.strftime('%d.%m.%Y')} | {p}",
            f"ABC_MI: A=золото (топ 70% MI), B=синій, C=сірий. ⚑ = наявність <{params['low_avail']}%.", COLS2)
    for ri, r in enumerate(sorted(results, key=lambda x: x['days_left'])):
        row=ri+4; dl=r['days_left']; low=r.get('low_avail',False)
        _,bg=ST.get(r['status'],(C['gray'],C['gray_l']))
        if low: bg=C['low_l']
        elif r['status']=='Норма': bg=C['row1'] if ri%2==0 else C['row2']
        tr=r.get('trend','—'); mi=r.get('mi_day'); abc_mi=r.get('abc_mi','?'); sd_=r.get('safety_disc',0)
        rec60_str=f"{r['rec_60']} шт ({sd_}д)" if r['rec_60']>0 else "—"
        vals=[r['sku'],r['name'],r['abc'],abc_mi,r['avg_margin'],mi,r['avail_pct'],
              tr,r['avg_day'],r['stock'],r['transit'],r.get('zero_date'),
              dl if dl<999 else None,r['rec'],rec60_str]
        for col,val in enumerate(vals,1):
            c=ws2.cell(row=row,column=col,value=val); c.fill=fl(bg); c.border=tb()
            if col==1:
                c.value=("⚑ " if low else "")+str(r['sku'])
                c.font=Font(name="Arial",bold=True,size=10,color="880E4F" if low else "000000"); c.alignment=la()
            elif col==2: c.font=cf(sz=9); c.alignment=la()
            elif col==3: c.font=Font(name="Arial",bold=True,size=11,color=AB.get(r['abc'],C['gray'])); c.alignment=ca()
            elif col==4: c.font=Font(name="Arial",bold=True,size=11,color=AB_MI.get(abc_mi,C['gray'])); c.alignment=ca()
            elif col==5:
                if val: c.number_format='0.0"%"'
                c.font=Font(name="Arial",bold=True,size=10,color=C['green'] if (val or 0)>=params['mg_a'] else C['blue']); c.alignment=ca()
            elif col==6:
                if val: c.number_format='#,##0.00'
                c.font=cf(sz=10,color=C['gold'] if (val or 0)>=(meta.get('ta') or 0) else C['blue']); c.alignment=ca()
            elif col==7:
                if val: c.number_format='0"%"'
                bc=C['red'] if (val or 0)<params['low_avail'] else C['amber'] if (val or 0)<50 else C['green']
                c.font=Font(name="Arial",bold=low,size=10,color=bc); c.alignment=ca()
            elif col==8: c.font=Font(name="Arial",bold=True,size=10,color=TR.get(tr,C['gray'])); c.alignment=ca()
            elif col==9: c.number_format='0.0000'; c.font=Font(name="Arial",bold=True,size=10,color=C['blue']); c.alignment=ca()
            elif col in(10,11): c.font=cf(sz=10); c.alignment=ca()
            elif col==12:
                if val:
                    fc2,_=ST.get(r['status'],(C['gray'],C['gray_l']))
                    c.font=Font(name="Arial",bold=True,size=10,color=fc2)
                else: c.value="∞"; c.font=cf(sz=10,color="AAAAAA")
                c.alignment=ca()
            elif col==13:
                fc2,_=ST.get(r['status'],(C['gray'],C['gray_l']))
                c.font=Font(name="Arial",bold=True,size=11,color=fc2); c.alignment=ca()
            elif col==14:
                if val and val>0: c.font=Font(name="Arial",bold=True,size=11,color=C['main'])
                else: c.value="—"; c.font=cf(sz=10,color="BBBBBB")
                c.alignment=ca()
            else: c.font=cf(sz=10); c.alignment=ca()
    ws2.freeze_panes="A4"; ws2.auto_filter.ref=f"A3:{get_column_letter(len(COLS2))}{len(results)+3}"

    # Sheet 3 — Ручне рішення
    ws3 = wb.create_sheet("Ручне рішення")
    low_rows = sorted([r for r in results if r.get('low_avail')], key=lambda x: x['avail_pct'])
    COLS3 = [("SKU",13),("Назва",42),("ABC",7),("Маржа %",10),
             ("Наявн %",9),("Тренд",11),("Залишок",8),("Дата нуля",11)]
    ws_init(ws3, f"Ручне рішення — наявність <{params['low_avail']}% | {len(low_rows)} SKU",
            "Прогноз ненадійний. Вирішіть вручну.", COLS3, "880E4F", "FCE4EC")
    for ri,r in enumerate(low_rows):
        row=ri+4; bg="FCE4EC" if ri%2==0 else C['white']; dl=r['days_left']; tr=r.get('trend','—')
        for col,val in enumerate([r['sku'],r['name'],r['abc'],r['avg_margin'],
                                   r['avail_pct'],tr,r['stock'],r.get('zero_date')],1):
            c=ws3.cell(row=row,column=col,value=val); c.fill=fl(bg); c.border=tb()
            if col==1: c.font=cf(bold=True,sz=10); c.alignment=la()
            elif col==2: c.font=cf(sz=9); c.alignment=la()
            elif col==3: c.font=Font(name="Arial",bold=True,size=11,color=AB.get(r['abc'],C['gray'])); c.alignment=ca()
            elif col==4:
                if val: c.number_format='0.0"%"'
                c.font=Font(name="Arial",bold=True,size=10,color=C['blue']); c.alignment=ca()
            elif col==5:
                if val: c.number_format='0"%"'
                c.font=Font(name="Arial",bold=True,size=10,color=C['red']); c.alignment=ca()
            elif col==6: c.font=Font(name="Arial",bold=True,size=10,color=TR.get(tr,C['gray'])); c.alignment=ca()
            elif col==7: c.font=cf(sz=10); c.alignment=ca()
            elif col==8:
                if val:
                    c.font=Font(name="Arial",size=10,color=C['red'] if dl<5 else C['amber'] if dl<15 else "000000")
                else: c.value="∞"; c.font=cf(sz=10,color="AAAAAA")
                c.alignment=ca()

    # Sheet 4 — Разовий попит
    ws4 = wb.create_sheet("Разовий попит")
    COLS4 = [("SKU",13),("Назва",42),("Маржа %",10),("Чист. міс.",12),("Продано",12),("Причина",28)]
    ws_init(ws4, f"Разовий попит — {len(sporadic)} SKU",
            f"< {params['min_months']} міс. АБО < {params['min_qty']} шт — виключені з замовлень.",
            COLS4, "BF360C", "FBE9E7")
    for ri,r in enumerate(sorted(sporadic,key=lambda x:-(x.get('avg_margin') or 0))[:300]):
        row=ri+4; bg=C['orange_l'] if ri%2==0 else C['white']
        cmo=sum(1 for lbl in (data['meta']['months'] if 'meta' in data else []))
        ts=r.get('total_clean_sold', 0)
        for col,val in enumerate([r['sku'],r['name'],r.get('avg_margin'),
                                   r.get('clean_months',0),ts,r.get('sporadic_reason','')],1):
            c=ws4.cell(row=row,column=col,value=val); c.fill=fl(bg); c.border=tb()
            if col==1: c.font=cf(bold=True,sz=10); c.alignment=la()
            elif col==2: c.font=cf(sz=9); c.alignment=la()
            elif col==3:
                if val: c.number_format='0.0"%"'
                c.font=cf(sz=10,color=C['blue']); c.alignment=ca()
            elif col==6: c.font=Font(name="Arial",size=10,italic=True,color=C['orange']); c.alignment=la()
            else: c.font=cf(sz=10); c.alignment=ca()

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ── UI ───────────────────────────────────────
st.title("📦 ProcureAI — Аналіз закупівель")
st.caption("Файл 1: Шаблон (продажі + наявність + залишки) | Файл 2: Ціни iHerb")

with st.sidebar:
    st.header("⚙️ Параметри")
    with st.expander("💰 Маржа та ABC", expanded=True):
        mg_min = st.slider("Мінімальна маржа (%)", 5, 50, 35)
        mg_a   = st.slider("A-клас: маржа ≥ (%)", mg_min, 70, max(mg_min+10, 50))
        a_mult = st.slider("Бонус запасу A-класу (×)", 1.0, 2.0, 1.3, 0.05)
    with st.expander("📅 Замовлення", expanded=True):
        lead    = st.slider("Lead time (дні)", 1, 30, 14)
        safety  = st.slider("Страховий запас (дні)", 7, 60, 30)
        st.markdown("**Знижки постачальника:**")
        disc_thr  = st.slider("Поріг знижки 1 (%)", 5, 60, 10, help="Знижка ≥ цього % → запас на N днів")
        safety60  = st.slider("Запас при знижці 1 (дні)", 30, 90, 60)
        disc_thr2 = st.slider("Поріг знижки 2 (%)", 15, 70, 40, help="Знижка ≥ цього % → більший запас")
        safety90  = st.slider("Запас при знижці 2 (дні)", 60, 180, 90)
        if disc_thr2 <= disc_thr:   disc_thr2 = disc_thr + 1
        if safety90  <= safety60:   safety90  = safety60 + 1
    with st.expander("📊 Попит", expanded=True):
        min_months = st.slider("Мін. місяців з продажами", 2, 9, 6)
        min_qty    = st.slider("Мін. продано штук", 1, 30, 12)
        low_avail  = st.slider("Поріг низької наявності (%)", 10, 40, 20)
    with st.expander("🔢 Коефіцієнти", expanded=False):
        avail_alpha = st.slider("Коефіцієнт наявності α", 0.3, 1.5, 0.7, 0.05,
                                help="avail_K = (наявн%)^α")
        lambda_val  = st.slider("Часові ваги λ", 0.05, 0.5, 0.25, 0.05,
                                help="Більше λ = більша вага останніх місяців")

params = dict(mg_min=mg_min, mg_a=mg_a, a_mult=a_mult,
              lead=lead, safety=safety, safety60=safety60, safety90=safety90,
              disc_thr=disc_thr, disc_thr2=disc_thr2,
              min_months=min_months, min_qty=min_qty, low_avail=low_avail,
              avail_alpha=avail_alpha, lambda_val=lambda_val)

# ── Завантаження файлів ──
col1, col2 = st.columns(2)
with col1:
    f_template = st.file_uploader(
        "📋 **Файл 1 — Шаблон**",
        type=["xlsx","xls"], key="template",
        help="Вкладки: 'продажі дані', 'наявність на складі', 'Залишки'")
with col2:
    f_prices = st.file_uploader(
        "💰 **Файл 2 — Ціни iHerb**",
        type=["xlsx","xls"], key="prices",
        help="iHerb Catalog з колонками: Артикул, Цена, Наличие, Скидка")

if f_template:
    with st.spinner("Читаємо шаблон..."):
        wb_t = load_wb(f_template.read())
        sku_data, months_labels, stock_map, avail_map, errors = parse_template(wb_t)

    if errors:
        for e in errors: st.error(e)
        st.stop()

    price_map = {}
    if f_prices:
        with st.spinner("Читаємо ціни..."):
            wb_p = load_wb(f_prices.read())
            price_map = parse_prices(wb_p)
        st.success(f"✅ Ціни завантажено: {len(price_map)} SKU, "
                   f"зі знижкою: {sum(1 for v in price_map.values() if v[2]>0)}")
    else:
        st.info("💡 Файл цін не завантажено — колонка знижки буде порожньою")

    # Зведення по вкладках
    with st.expander("🔍 Статистика завантажених даних", expanded=False):
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("SKU в продажах", len(sku_data))
        c2.metric("SKU в наявності", len(stock_map))
        c3.metric("SKU в залишках", len(avail_map))
        c4.metric("Місяців даних", len(months_labels) if months_labels else 0)
        st.write(f"**Місяці:** {', '.join(months_labels) if months_labels else '—'}")

    with st.spinner("Аналізуємо..."):
        data = run_analysis(sku_data, months_labels, stock_map, avail_map, price_map, params)

    meta     = data['meta']
    regular  = data['regular']
    sporadic = data['sporadic']
    excl_mg  = data['excl_mg']
    total    = len(regular)+len(sporadic)+len(excl_mg)

    # KPI
    st.subheader("📊 Зведення")
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    c1.metric("Всього SKU", total)
    c2.metric("Постійний попит", len(regular))
    c3.metric("Разовий попит", len(sporadic))
    c4.metric("Критично (<5дн)", sum(1 for r in regular if r['status']=='Критично'))
    c5.metric("До замовлення", sum(1 for r in regular if r['rec']>0))
    c6.metric("Season K", f"{meta['season_K']:.3f}")

    st.subheader("📅 Дата нуля залишків")
    buckets = [("🔴 Вже 0",0,1),("🔴 1-7 дн",1,8),("🟡 8-14 дн",8,15),
               ("🟡 15-30 дн",15,31),("🟢 31-60 дн",31,61),("🔵 >60 дн",61,999)]
    cols_b = st.columns(len(buckets))
    for i,(lbl,lo,hi) in enumerate(buckets):
        cols_b[i].metric(lbl, sum(1 for r in regular if lo<=r['days_left']<hi))

    tab1,tab2,tab3,tab4 = st.tabs(["🛒 Замовлення","📋 Аналіз SKU","⚠️ Ручне рішення","📦 Разовий попит"])

    with tab1:
        order = sorted([r for r in regular if r['rec']>0 or r['rec_60']>0],
                       key=lambda x: x['days_left'])
        if order:
            df = pd.DataFrame([{
                'SKU':       ('⚑ ' if r['low_avail'] else '')+r['sku'],
                'Назва':     r['name'][:50],
                'ABC':       r['abc'], 'ABC_MI': r.get('abc_mi','?'),
                'Маржа %':   r['avg_margin'],
                'MI/день':   r.get('mi_day'),
                'Залишок':   r['eff_stock'],
                'Дата нуля': r.get('zero_date','∞'),
                'Дні до 0':  r['days_left'] if r['days_left']<999 else '∞',
                'Тренд':     r['trend'],
                'Замовити 14+30': r['rec'],
                'Замовити (знижка)': (f"{r['rec_60']} шт ({r.get('safety_disc',0)}д)"
                                      if r.get('rec_60',0)>0 else '—'),
                'Знижка %':  r.get('price_disc',0) or '—',
            } for r in order])
            st.dataframe(df, use_container_width=True, height=500)
            disc1 = [r for r in order if r.get('rec_60',0)>0 and not r.get('use_90')]
            disc2 = [r for r in order if r.get('rec_60',0)>0 and r.get('use_90')]
            st.caption(
                f"Всього: {len(order)} SKU | стандарт: {sum(r['rec'] for r in order)} шт | "
                f"знижка 1 ({params['safety60']}д): {sum(r.get('rec_60',0) for r in disc1)} шт / {len(disc1)} SKU | "
                f"знижка 2 ({params.get('safety90',90)}д): {sum(r.get('rec_60',0) for r in disc2)} шт / {len(disc2)} SKU")
        else:
            st.success("Всі залишки в нормі — замовлення не потрібні")

    with tab2:
        df2 = pd.DataFrame([{
            'SKU':       ('⚑ ' if r['low_avail'] else '')+r['sku'],
            'Назва':     r['name'][:50],
            'ABC':       r['abc'], 'ABC_MI': r.get('abc_mi','?'),
            'Маржа %':   r['avg_margin'],
            'MI/день':   r.get('mi_day'),
            'Наявн %':   r['avail_pct'],
            'Тренд':     r['trend'],
            'avg/день':  r['avg_day'],
            'Залишок':   r['stock'],
            'Транзит':   r['transit'],
            'Дата нуля': r.get('zero_date','∞'),
            'Дні до 0':  r['days_left'] if r['days_left']<999 else '∞',
            'Статус':    r['status'],
            'Замовити':  r['rec'],
        } for r in sorted(regular, key=lambda x: x['days_left'])])
        st.dataframe(df2, use_container_width=True, height=500)

    with tab3:
        low_rows = sorted([r for r in regular if r.get('low_avail')], key=lambda x: x['avail_pct'])
        if low_rows:
            df3 = pd.DataFrame([{'SKU':r['sku'],'Назва':r['name'][:50],'ABC':r['abc'],
                                  'Маржа %':r['avg_margin'],'Наявн %':r['avail_pct'],
                                  'Тренд':r['trend'],'Дата нуля':r.get('zero_date','∞')}
                                 for r in low_rows])
            st.warning(f"⚑ {len(low_rows)} SKU мали наявність <{params['low_avail']}%")
            st.dataframe(df3, use_container_width=True, height=400)
        else:
            st.success("Немає SKU з низькою наявністю")

    with tab4:
        if sporadic:
            df4 = pd.DataFrame([{'SKU':r['sku'],'Назва':r['name'][:50],
                                  'Маржа %':r.get('avg_margin'),'Причина':r.get('sporadic_reason','')}
                                 for r in sorted(sporadic, key=lambda x:-(x.get('avg_margin') or 0))[:300]])
            st.info(f"🔵 {len(sporadic)} SKU з нерегулярним попитом")
            st.dataframe(df4, use_container_width=True, height=400)
        else:
            st.success("Всі SKU мають постійний попит")

    st.divider()
    with st.spinner("Генеруємо Excel..."):
        excel_buf = gen_excel(data, params)
    st.download_button(
        "📥 Завантажити Excel-звіт (4 вкладки)",
        data=excel_buf,
        file_name=f"ProcureAI_{date.today().strftime('%d%m%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)

else:
    st.info("👆 Завантажте Файл 1 (Шаблон) щоб почати аналіз")
    with st.expander("📖 Структура файлів"):
        st.markdown("""
**Файл 1 — Шаблон** (оновлюєте щодня):
| Вкладка | Структура |
|---|---|
| `продажі дані` | Блоки по 5 колонок на місяць: Назва | Артикул | Кількість | Рентаб% | _ |
| `наявність на складі` | Номенклатура | Артикул | Залишок | Замовлено у постачальників |
| `Залишки` | Блоки по 4 колонки на місяць: Артикул | Назва | Дні на складі | _ |

**Файл 2 — Ціни iHerb** (оновлюєте за потреби):
Артикул | Название | Цена | Старая цена | Наличие | Скидка | ...
        """)
